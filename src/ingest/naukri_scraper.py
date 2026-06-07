from __future__ import annotations

import re
import time
from datetime import datetime
from pathlib import Path
from urllib.parse import quote_plus, urljoin, urlparse

from openpyxl import Workbook, load_workbook
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright


BASE_URL = "https://www.naukri.com"
SHEET_NAME = "Job Links"
HEADERS = ["Technology", "Job Title", "Company", "Posted", "Email", "Contact Number", "Job Link", "Collected At"]
EMAIL_PATTERN = re.compile(r"\b[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}\b", re.IGNORECASE)
PHONE_PATTERN = re.compile(
    r"(?:\+91[\s-]?)?(?:0[\s-]?)?[6-9]\d{2}[\s-]?\d{3}[\s-]?\d{4}\b"
)


def normalize_url(url: str) -> str:
    parsed = urlparse(url)
    clean = parsed._replace(query="", fragment="")
    return clean.geturl().rstrip("/")


def is_job_link(href: str) -> bool:
    if not href:
        return False
    parsed = urlparse(href)
    host = parsed.netloc.lower()
    path = parsed.path.lower()
    return "naukri.com" in host and (
        "/job-listings-" in path
        or "/job-listings/" in path
        or re.search(r"-\d{8,}$", path) is not None
    )


def parse_age_to_hours(text: str) -> float | None:
    text = " ".join(text.lower().split())
    if not text:
        return None
    if "just now" in text or "few minutes" in text or "today" in text:
        return 0
    if "yesterday" in text:
        return 24

    match = re.search(r"(\d+)\s*(minute|min|mins|minutes)\b", text)
    if match:
        return int(match.group(1)) / 60

    match = re.search(r"(\d+)\s*(hour|hr|hrs|hours)\b", text)
    if match:
        return int(match.group(1))

    match = re.search(r"(\d+)\s*(day|days)\b", text)
    if match:
        return int(match.group(1)) * 24

    match = re.search(r"(\d+)\s*(week|weeks|month|months|year|years)\b", text)
    if match:
        return 9999

    return None


def ensure_headers(sheet) -> None:
    existing = [cell.value for cell in sheet[1]]
    if existing == HEADERS:
        return

    rows = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {}
        for index, header in enumerate(existing):
            if header:
                row_data[str(header)] = values[index] if index < len(values) else ""
        rows.append(row_data)

    if sheet.max_row:
        sheet.delete_rows(1, sheet.max_row)

    sheet.append(HEADERS)
    for row_data in rows:
        sheet.append([row_data.get(header, "") for header in HEADERS])


def ensure_workbook(path: Path):
    if path.exists():
        workbook = load_workbook(path)
        if SHEET_NAME not in workbook.sheetnames:
            sheet = workbook.create_sheet(SHEET_NAME)
            sheet.append(HEADERS)
            workbook.save(path)
        else:
            sheet = workbook[SHEET_NAME]
            ensure_headers(sheet)
            workbook.save(path)
        return workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.append(HEADERS)
    workbook.save(path)
    return workbook


def header_map(sheet) -> dict[str, int]:
    return {str(cell.value): index for index, cell in enumerate(sheet[1], start=1) if cell.value}


def existing_links(sheet) -> set[str]:
    links = set()
    headers = header_map(sheet)
    link_column = headers.get("Job Link")
    if not link_column:
        return links

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if len(row) >= link_column and row[link_column - 1]:
            links.add(normalize_url(str(row[link_column - 1])))
    return links


def search_url(keyword: str, page: int) -> str:
    query = quote_plus(keyword)
    if page <= 1:
        return f"{BASE_URL}/{query}-jobs?k={query}"
    return f"{BASE_URL}/{query}-jobs-{page}?k={query}"


def extract_contact_details(text: str) -> tuple[str, str]:
    emails = sorted(set(EMAIL_PATTERN.findall(text)))
    phones = sorted(set(match.group(0).strip() for match in PHONE_PATTERN.finditer(text)))
    return ", ".join(emails), ", ".join(phones)


def fetch_contact_details(browser, link: str) -> tuple[str, str]:
    detail_page = browser.new_page()
    try:
        detail_page.goto(link, wait_until="domcontentloaded", timeout=45000)
        detail_page.wait_for_timeout(2500)
        text = detail_page.locator("body").inner_text(timeout=5000)
        return extract_contact_details(text)
    except Exception:
        return "", ""
    finally:
        detail_page.close()


def extract_jobs(page, max_age_hours: int):
    jobs = []
    cards = page.locator("article, .srp-jobtuple-wrapper, .jobTuple, div:has(a[href*='job-listings'])")
    count = min(cards.count(), 80)

    for index in range(count):
        card = cards.nth(index)
        try:
            anchor = card.locator("a[href*='job-listings']").first
            href = anchor.get_attribute("href", timeout=1000)
            if not href:
                continue

            link = normalize_url(urljoin(BASE_URL, href))
            if not is_job_link(link):
                continue

            title = anchor.inner_text(timeout=1000).strip()
            card_text = card.inner_text(timeout=1000)
            lower_text = card_text.lower()
            email, contact_number = extract_contact_details(card_text)

            posted_text = ""
            age_hours = None
            for line in card_text.splitlines():
                if any(word in line.lower() for word in ["ago", "today", "yesterday", "just now"]):
                    posted_text = line.strip()
                    age_hours = parse_age_to_hours(posted_text)
                    break

            if age_hours is None:
                age_hours = parse_age_to_hours(lower_text)
                if age_hours is None:
                    continue

            if age_hours > max_age_hours:
                continue

            company = ""
            company_locator = card.locator(".comp-name, .companyName, a[href*='company']")
            if company_locator.count() > 0:
                company = company_locator.first.inner_text(timeout=1000).strip()

            jobs.append(
                {
                    "title": title,
                    "company": company,
                    "posted": posted_text,
                    "email": email,
                    "contact_number": contact_number,
                    "link": link,
                }
            )
        except PlaywrightTimeoutError:
            continue
        except Exception:
            continue

    if jobs:
        return jobs

    anchors = page.locator("a[href]")
    count = min(anchors.count(), 300)
    for index in range(count):
        try:
            anchor = anchors.nth(index)
            href = anchor.get_attribute("href", timeout=1000)
            link = normalize_url(urljoin(BASE_URL, href or ""))
            if is_job_link(link):
                jobs.append(
                    {
                        "title": anchor.inner_text(timeout=1000).strip(),
                        "company": "",
                        "posted": "",
                        "email": "",
                        "contact_number": "",
                        "link": link,
                    }
                )
        except Exception:
            continue
    return jobs


def wait_for_manual_login(page):
    page.goto(BASE_URL, wait_until="domcontentloaded")
    print("\nA browser window is open.")
    print("Log in to Naukri manually if you are not already logged in.")
    print("When login is complete, return here and press Enter.")
    input()
    try:
        page.wait_for_load_state("networkidle", timeout=15000)
    except Exception:
        pass


def goto_with_retry(page, url: str, timeout: int = 60000):
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=timeout)
    except Exception as exc:
        message = str(exc).lower()
        if "interrupted by another navigation" not in message and "navigation" not in message:
            raise

        page.wait_for_load_state("networkidle", timeout=15000)
        page.goto(url, wait_until="domcontentloaded", timeout=timeout)


def scrape_job_links(
    *,
    keyword: str,
    output_path: Path,
    profile_dir: Path,
    pages: int,
    max_age_hours: int = 24,
    delay: float = 2.0,
    login: bool = False,
    skip_contact_details: bool = False,
) -> dict[str, int]:
    output_path = Path(output_path).resolve()
    profile_dir = Path(profile_dir).resolve()
    profile_dir.mkdir(parents=True, exist_ok=True)

    workbook = ensure_workbook(output_path)
    sheet = workbook[SHEET_NAME]
    seen = existing_links(sheet)

    added = 0
    scanned = 0
    collected_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch_persistent_context(
            user_data_dir=str(profile_dir),
            headless=False,
            viewport={"width": 1366, "height": 900},
        )
        page = browser.new_page()

        if login:
            wait_for_manual_login(page)

        for page_number in range(1, pages + 1):
            url = search_url(keyword, page_number)
            print(f"Scanning page {page_number}: {url}")
            goto_with_retry(page, url)
            page.wait_for_timeout(3500)

            jobs = extract_jobs(page, max_age_hours)
            scanned += len(jobs)

            for job in jobs:
                normalized = normalize_url(job["link"])
                if normalized in seen:
                    continue

                if not skip_contact_details and (not job["email"] or not job["contact_number"]):
                    email, contact_number = fetch_contact_details(browser, job["link"])
                    job["email"] = job["email"] or email
                    job["contact_number"] = job["contact_number"] or contact_number

                sheet.append(
                    [
                        keyword,
                        job["title"],
                        job["company"],
                        job["posted"],
                        job["email"],
                        job["contact_number"],
                        job["link"],
                        collected_at,
                    ]
                )
                seen.add(normalized)
                added += 1

            workbook.save(output_path)

            if not jobs:
                print("No matching jobs found on this page. Stopping search.")
                break
            time.sleep(delay)

        browser.close()

    print("\nDone.")
    print(f"Scanned matching recent jobs: {scanned}")
    print(f"New links added: {added}")
    print(f"Excel file: {output_path}")
    return {"scanned": scanned, "added": added}
